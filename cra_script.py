import re
import sys
import requests
import tkinter as tk
from openpyxl import Workbook
from bs4 import BeautifulSoup
from openpyxl.utils import column_index_from_string, get_column_letter


class crawler:
    # 定义要爬取的 JSP 网页的 URL
    staic_url = "http://xsgl.7i5q.cas.scut.edu.cn"
    url = "http://xsgl.7i5q.cas.scut.edu.cn/sms2/student/evaluation/intellectualList.jsp"

    # 创建一个新的工作簿
    workbook = Workbook()

    # 获取活动的工作表
    sheet = workbook.active

    sheet.cell(row=1, column=1).value = "姓名"

    # 用于存储每个课程对应的列索引
    course_columns = {}  # 记录课程与索引之间的关系
    name_rows = {}  # 记录姓名与索引之间的关系
    column_index = 2
    row_index = 1

    def __init__(self, screen):
        self.screen = screen

    # gui显示文本
    def areaInsert(self, str):
        self.screen.insert(tk.END, str + "\n")
        self.screen.see(tk.END)
        print(str)

    # 查找体育课的位置
    def serach_physical_index(self, semester):
        tar_str = "体育" + str(semester)
        indexs = []
        for column in self.sheet.iter_cols():
            if column[0].value.startswith(tar_str):
                indexs.append(column[0].column)
        return indexs

    # 处理每个成绩项的函数
    def processing_grade_item(self, tr):
        tds = tr.find_all("td")
        # 判断是否为必修课
        if tds[3].text.startswith("必修课"):
            course_name = tds[0].text.strip()
            pattern = re.compile(r'[^\u4e00-\u9fa5]')
            if not pattern.match(tds[1].text.strip()[:2]):
                grade = tds[1].text.split('（')[1].split('）')[0]
            else:
                grade = tds[1].text.strip()
            credit = tds[2].text.strip()
            return course_name, grade, credit
        else:
            return None, None, None

    # 添加成绩数据到表格的操作
    def add_data_to_sheet(self, r_index, course_name, grade, credit, semester):
        if not course_name == None:
            if "英语" in course_name:
                if "一" in course_name:
                    course_name = "英语(一)"
                else:
                    course_name = "英语(二)"

            if not course_name in self.course_columns:
                # 确保添加的成绩在体育课成绩之前
                if not len(self.serach_physical_index(semester)) == 0:
                    phy_index = column_index_from_string(self.serach_physical_index(semester)[0])
                    self.sheet.insert_cols(phy_index)
                    self.sheet.insert_cols(phy_index + 1)
                    self.course_columns[course_name] = phy_index
                    self.sheet.cell(row=1, column=phy_index).value = course_name
                    self.sheet.cell(row=1, column=phy_index + 1).value = "学分"
                    self.column_index += 2
                    self.course_columns["体育" + str(semester)] = phy_index + 2
                else:
                    self.course_columns[course_name] = self.column_index
                    self.sheet.cell(row=1, column=self.column_index).value = course_name
                    self.sheet.cell(row=1, column=self.column_index + 1).value = "学分"
                    self.column_index += 2

            index = self.course_columns[course_name]
            self.sheet.cell(row=r_index, column=index).value = grade
            self.sheet.cell(row=r_index, column=index + 1).value = credit

    # 处理一个学期的成绩，添加到excel表格
    def processing_grade(self, r_index, semester_items, semester):
        for i in range(1, len(semester_items)):
            course_name, grade, credit = self.processing_grade_item(semester_items[i])
            # 添加数据
            self.add_data_to_sheet(r_index, course_name, grade, credit, semester)

    # 获取体育成绩
    def get_physical_grade(self, yearId, d_src, headers, payload, student_name):
        physical_url = "/sms2/student/module/evaluation/studentGymDetail.jsp?" + d_src.split('?')[1]
        physical_response = requests.post(self.staic_url + physical_url, headers=headers, data=payload)

        # 检查响应状态码
        if physical_response.status_code == 200:
            self.areaInsert(yearId + ":" + student_name + "文体页爬取成功")
        else:
            self.areaInsert(yearId + ":" + student_name + "文体页爬取失败")
            sys.exit()

        # 使用 BeautifulSoup 解析网页内容
        physical_soup = BeautifulSoup(physical_response.content, "html.parser")

        try:
            # 成绩项
            physical_trs = physical_soup.find_all("tr")
            grade_trs = []
            grade_trs.append(physical_trs[8])
            grade_trs.append(physical_trs[9])
            for tr in grade_trs:
                tds = tr.find_all("td")
                if tds[1].text.strip() == "第一学期":
                    first_grade = tds[2].text.strip()
                    first_credit = 1.0
                else:
                    second_grade = tds[2].text.strip()
                    second_credit = 1.0
        except Exception as e:
            first_grade = 0.0
            first_credit = 0.0
            second_grade = 0.0
            second_credit = 0.0

        return first_grade, second_grade, first_credit, second_credit

    # student——process
    def student_process(self, years_index, f_seme, s_seme, first_grade, second_grade, first_credit, second_credit,
                        student_r_index):
        # 第一学期文化课成绩
        self.processing_grade(student_r_index, f_seme, 2 * years_index - 1)
        # 第一学期体育课成绩
        self.add_data_to_sheet(student_r_index, "体育" + str(2 * years_index - 1), first_grade, first_credit,
                               2 * years_index - 1)
        # 第二学期文化课成绩
        self.processing_grade(student_r_index, s_seme, 2 * years_index)
        # 第二学期体育课成绩
        self.add_data_to_sheet(student_r_index, "体育" + str(2 * years_index), second_grade, second_credit,
                               2 * years_index)

    # 主过程
    def Crawling_def(self, cookie_string, classYearIds):
        # 将 Cookie 添加到请求头中
        headers = {
            "Cookie": cookie_string
        }

        # 学年
        years_index = 1

        # 遍历每个学年
        for yearId in classYearIds:
            # 构造POST请求的参数
            payload = {
                'classYearId': yearId,
            }

            # 发送带有 Cookie 的请求
            response = requests.post(self.url, headers=headers, data=payload)

            # 检查响应状态码
            if response.status_code == 200:
                print("爬取成功")
                self.areaInsert("yearId:" + yearId + "爬取成功")
            else:
                self.areaInsert("yearId:" + yearId + "爬取失败")
                sys.exit()

            # 使用 BeautifulSoup 解析网页内容
            soup = BeautifulSoup(response.content, "html.parser")

            # 获取网页标题
            title = soup.title.string
            self.areaInsert("网页标题:" + title)

            # 找到所有的 <tr> 标签
            table_tags = soup.find_all("table")
            tmp = table_tags[0].find_all("tr")
            tmp1 = tmp[3].find_all("table")
            tr_tags = tmp1[1].find_all("tr")

            # 统计记录项项数
            num_of_item = re.findall(r'\d+', tmp1[2].find_all("td")[1].text)[0]

            """
            处理每个学生的具体成绩项
            """
            # 跳转详情页
            for j in range(1, int(num_of_item) + 1):
                student_name = tr_tags[j].find_all("a")[1].text.strip()
                detail_src = tr_tags[j].find_all("a")[2].get("href")

                # 发送带有 Cookie 的请求
                detail_response = requests.get(self.staic_url + detail_src, headers=headers)

                # 检查响应状态码
                if detail_response.status_code == 200:
                    self.areaInsert(yearId + ":" + student_name + "详情页爬取成功")
                else:
                    self.areaInsert(yearId + ":" + student_name + "详情页爬取失败")
                    continue

                # 使用 BeautifulSoup 解析网页内容
                detail_soup = BeautifulSoup(detail_response.content, "html.parser")

                # 没有权限访问
                if detail_soup.find_all("body")[0].text.strip().startswith("对不起"):
                    continue

                # 获得学生所在的行索引
                if not student_name in self.name_rows:
                    self.row_index = self.row_index + 1
                    self.sheet.cell(row=self.row_index, column=1).value = student_name
                    student_row_index = self.row_index
                    self.name_rows[student_name] = self.row_index
                else:
                    student_row_index = self.name_rows[student_name]

                # 成绩项<tr>标签整学年成绩
                detail_table_tags = detail_soup.find_all("table")
                table_trs = detail_table_tags[0].find_all("tr")
                table_tr_tables = table_trs[5].find_all("table")

                # 两学期成绩单
                first_semester_items = table_tr_tables[1].find_all("tr")
                second_semester_items = table_tr_tables[4].find_all("tr")

                first_grade, second_grade, first_credit, second_credit = self.get_physical_grade(yearId, detail_src,
                                                                                                 headers,
                                                                                                 payload, student_name)
                self.student_process(years_index, first_semester_items, second_semester_items, first_grade,
                                     second_grade,
                                     first_credit, second_credit, student_row_index)
            years_index += 1

        self.workbook.save("data.xlsx")
        self.areaInsert("源数据爬取完成")
