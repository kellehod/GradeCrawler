import tkinter as tk
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font
from openpyxl.utils import column_index_from_string, get_column_letter


class decorator:
    def __init__(self, screen):
        self.screen = screen

    def areaInsert(self, str):
        self.screen.insert(tk.END, str + "\n")
        self.screen.see(tk.END)
        print(str)

    def traverse_search(self, sheet, indexs, tar_str):
        for column in sheet.iter_cols():
            if column[0].value.startswith(tar_str):
                indexs.append(column[0].column)

    # 计算学期加权总分和总学分数
    def calculate_total_data(self, sheet, semester, partition_index):
        start_index = partition_index[(semester - 1) * 2]
        end_index = partition_index[(semester - 1) * 2 + 1]
        total_courses = (end_index - start_index + 1) // 2
        semester_credit = 0.0
        for row_index in range(2, sheet.max_row + 1):
            total_score = 0.0
            total_credit = 0.0
            course = 0
            while course < total_courses:
                index = start_index + course * 2
                grade_value = sheet.cell(row=row_index, column=index).value
                if grade_value is not None:
                    grade = float(grade_value)
                else:
                    grade = 0.0
                credit_value = sheet.cell(row=row_index, column=index + 1).value
                if credit_value is not None:
                    credit = float(credit_value)
                else:
                    credit = 0.0
                total_credit += credit
                score = grade * credit
                total_score += score
                course += 1
            try:
                average_score = total_score / total_credit
            except Exception as e:
                average_score = 0
            sheet.cell(row=row_index, column=(2 + (semester - 1) * 3)).value = average_score
            sheet.cell(row=row_index, column=(2 + (semester - 1) * 3) + 1).value = total_score
        
    # 排序,指定需要排序的列
    def ranking(self, sheet, baseindex, offset):
        data_range = sheet[get_column_letter(baseindex)][1:]
        data = [cell.value for cell in data_range]

        # 使用sorted()函数按照从大到小的顺序排序数组，并保留原始索引
        sorted_data = sorted(enumerate(data), key=lambda x: x[1], reverse=True)

        ranked_data = []  # 用于存储排名的列表
        current_rank = 1  # 当前排名，默认为1

        for i, (index, value) in enumerate(sorted_data):
            if i > 0 and value < sorted_data[i - 1][1]:
                current_rank = i + 1  # 更新当前排名
            ranked_data.append((index, current_rank))

        for index, current_rank in ranked_data:
            sheet.cell(row=index + 2, column=baseindex + offset).value = current_rank

    # 总均分，排名
    def total_semester(self, sheet):
        sheet.insert_cols(2)
        sheet.insert_cols(3)
        sheet.cell(row=1, column=2).value = "总加权均分"
        sheet.cell(row=1, column=3).value = "总排名"
        for i in range(2, sheet.max_row + 1):
            totalScore = 0.0
            total_credits = 0.0
            for j in range(self.semester_count):
                score = float(sheet.cell(row=i, column=4 + j * 3).value)
                tol_score = float(sheet.cell(row=i, column=4 + j * 3 + 1).value)
                if not score == 0:
                    credit = tol_score / score
                else:
                    credit = 0.0
                total_credits += credit
                totalScore += tol_score
            try:
                avg_score = totalScore / total_credits
            except Exception as e:
                avg_score = 0
            sheet.cell(row=i, column=2).value = avg_score
        self.ranking(sheet, 2, 1)

    def decoration(self, ):
        # 学期数
        self.semester_count = 0
        # 在第一列后插入第一学期的空列
        self.insert_column_index = 2  # 在第一列后插入，索引从1开始
        # 遍历第一行，查找总共有多少门体育课
        self.target_column_index = []
        # 打开XLSX文件
        try:
            workbook = load_workbook('data.xlsx')
            sheet = workbook.active
        except Exception as e:
            self.areaInsert("未存在源文件")
            return

        self.traverse_search(sheet, self.target_column_index, "体育")

        # 为每个学期预留空列
        for num in range(len(self.target_column_index)):
            self.semester_count += 1
            for i in range(3):
                sheet.insert_cols(self.insert_column_index)
                self.insert_column_index += 1
            sheet.cell(row=1, column=self.insert_column_index - 3).value = "加权均分" + str(self.semester_count)
            sheet.cell(row=1, column=self.insert_column_index - 2).value = "加权总分" + str(self.semester_count)
            sheet.cell(row=1, column=self.insert_column_index - 1).value = "排名" + str(self.semester_count)

        partition_index = [2 + 3 * self.semester_count]
        for column in sheet.iter_cols():
            if column[0].value.startswith("体育"):
                partition_index.append(column_index_from_string(column[0].column) + 1)  # 学期最后一门课列索引
                partition_index.append(column_index_from_string(column[0].column) + 2)  # 学期开始一门课列索引

        for sem in range(1, self.semester_count + 1):
            self.calculate_total_data(sheet, sem, partition_index)
            self.ranking(sheet, (2 + (sem - 1) * 3), 2)

        self.total_semester(sheet)

        # 居中显示
        for row in range(1, sheet.max_row + 1):
            for column in range(1, sheet.max_column + 1):
                sheet.cell(row=row, column=column).alignment = Alignment(horizontal='center', vertical='center')

        # 创建填充样式
        fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

        # 创建字体样式
        bold_font = Font(bold=True)
        red_font = Font(color="FF0000")  # 红色的RGB值为 "FF0000"
        blue_font = Font(color="0000FF")  # 蓝色的RGB值为 "0000FF"

        # 创建边框样式
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # 第一行全加粗
        for column in sheet.iter_cols():
            column[0].font = bold_font

        # 均分, 排名
        for cell in sheet['B'][1:]:
            cell.font = red_font
            right_cell = cell.offset(row=0, column=1)
            right_cell.font = blue_font

        for semester in range(self.semester_count):
            for cell in sheet[get_column_letter(4 + semester * 3)][1:]:
                cell.font = red_font
                right_cell = cell.offset(row=0, column=2)
                right_cell.font = blue_font

        # 成绩区域样式调整
        credit_column_indexs = []
        self.traverse_search(sheet, credit_column_indexs, "学分")
        for column in credit_column_indexs:
            sheet.column_dimensions[column].width = 6
            for cell in sheet[column]:
                cell.fill = fill
                cell.border = border
                left_cell = cell.offset(row=0, column=-1)
                left_cell.font = bold_font

        # 保存修改后的文件
        workbook.save('target_data.xlsx')
        self.areaInsert("分析完成")
